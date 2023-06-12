import os
import git
import stat
import json
import requests
import openpyxl
import shutil
import pandas
from time import sleep
from datetime import datetime
from atlassian import Bitbucket


proxy_setting1 = "http://farft01:fardnc01@proxy-us.cnhind.com:8080"
proxy_setting2 = "https://farft01:fardnc01@proxy-us.cnhind.com:8080"
basic_auth = ("dylanrsmith", "ATBB8QqnhrB3vr8k8aMvq4hrJdUy56EB4A95")
bitbucket = Bitbucket(
    url="http://localhost:7990",
    username="dylanrsmith",
    password="ATBB8QqnhrB3vr8k8aMvq4hrJdUy56EB4A95",
)
bitbucket._session.proxies = {"http": proxy_setting1, "https": proxy_setting2}


# python-git
username = "dylanrsmith"
password = "ATBB8QqnhrB3vr8k8aMvq4hrJdUy56EB4A95"
repo = git.Repo()
repo_names = []  # repo ID BB
repo_staging_spots = []  # Local Download Path
repo_onedrive_spots = []  # OneDrive Repo Upload Path
plant_server_spots = []  # server path
plant_onedrive_spots = []  # onedrive path
triggers = []  # 'W' or 'X'
repository = []


# fargoengineering domain - Azure Active Directory
client_id = "5421bf2a-f352-4a3d-a647-0a8ded1e96d7"
client_secret = "~_C8Q~t8K-gFvszz7r5sRu-IlBWDnJ~Pa2W10aBu"
tenant = "b982fbe0-f860-4be8-b974-25005ff5aba4"
redirect_uri = "http://localhost:8080"
user_id = "Admin@fargoengineering.com"


def get_access_token(client_id, client_secret, tenant):
    token_url = f"https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token"
    token_data = {
        "client_id": client_id,
        "client_secret": client_secret,
        "scope": "https://graph.microsoft.com/.default",
        "grant_type": "client_credentials",
    }
    token_response = requests.post(token_url, data=token_data)
    access_token = token_response.json().get("access_token")
    return access_token


def get_folder_id(access_token, folder_path):
    headers = {
        "Authorization": f"Bearer {access_token}",
    }

    get_folder_url = (
        f"https://graph.microsoft.com/v1.0/users/{user_id}/drive/root:/{folder_path}"
    )

    response = requests.get(get_folder_url, headers=headers)

    if response.status_code == 200:
        print(f"Folder {folder_path} is found successfully")
        return response.json()["id"]
    else:
        print(
            f"Failed to find folder {folder_path}, status code: {response.status_code}"
        )
        if response.content:
            print(f"Error message: {response.json()}")
        return None


def download_folder_from_onedrive(
    access_token, onedrive_folder_path, local_folder_path
):
    get_children_url = f"https://graph.microsoft.com/v1.0/users/Admin@fargoengineering.com/drive/items/root:/{onedrive_folder_path}:/children"

    headers = {
        "Content-Type": "application/json",
        "Authorization": "Bearer " + access_token,
    }

    response = requests.get(get_children_url, headers=headers)

    if response.status_code in [200, 201, 202]:
        children = response.json()["value"]  # List of all children of the folder

        if not os.path.exists(local_folder_path):
            os.makedirs(
                local_folder_path
            )  # Create the local directory if it doesn't exist

        for child in children:
            if "folder" not in child:  # If the child is a file
                onedrive_file_path = f"{onedrive_folder_path}/{child['name']}"
                local_file_path = f"{local_folder_path}/{child['name']}"
                download_file_from_onedrive(
                    access_token, onedrive_file_path, local_file_path
                )
    else:
        print(
            f"Error while trying to get list of children, status code: {response.status_code}"
        )


def download_file_from_onedrive(access_token, onedrive_file_path, local_file_path):
    download_url = f"https://graph.microsoft.com/v1.0/users/Admin@fargoengineering.com/drive/items/root:/{onedrive_file_path}:/content"

    headers = {
        "Content-Type": "application/json",
        "Authorization": "Bearer " + access_token,
    }

    response = requests.get(download_url, headers=headers)

    if response.status_code in (200, 201, 202, 204):
        with open(local_file_path, "wb") as file:
            file.write(response.content)
        print("File downloaded successfully.")
    else:
        print("Error downloading file:", response.content)


def upload_file_to_onedrive(access_token, folder_id, file_path):
    headers = {"Authorization": "Bearer " + access_token}
    file_name = os.path.basename(file_path)
    file_content = open(file_path, "rb").read()

    upload_url = f"https://graph.microsoft.com/v1.0/users/{user_id}/drive/items/{folder_id}:/{file_name}:/content"
    response = requests.put(upload_url, headers=headers, data=file_content)

    if response.status_code in [200, 201, 202, 203]:
        print(f"File {file_name} is uploaded successfully")
        return json.loads(response.content)
    else:
        print(f"Failed to upload file {file_name}, status code: {response.status_code}")
        return None


def folder_exists_in_response(response, folder_name):
    for item in response.json()["value"]:
        if item["name"] == folder_name and item["folder"]:
            return True
    return False


def create_folder_in_onedrive(access_token, parent_id, folder_name):
    headers = {
        "Authorization": "Bearer " + access_token,
        "Content-Type": "application/json",
    }

    create_folder_url = f"https://graph.microsoft.com/v1.0/users/{user_id}/drive/items/{parent_id}/children"

    response = requests.get(
        create_folder_url, headers=headers
    )  # Check if folder exists

    if response.status_code == 200:
        folders = response.json()["value"]  # List of all child folders
        if any(
            folder["name"] == folder_name for folder in folders
        ):  # Check if desired folder exists in the list
            print(f"Folder {folder_name} already exists.")
            return [folder for folder in folders if folder["name"] == folder_name][
                0
            ]  # return the folder's details
    else:
        print(
            f"Error while trying to get list of folders, status code: {response.status_code}"
        )
        return None

    # If we reach here, it means the folder doesn't exist. So, let's create it
    folder_metadata = {
        "name": folder_name,
        "folder": {},
        "@microsoft.graph.conflictBehavior": "rename",
    }

    response = requests.post(create_folder_url, headers=headers, json=folder_metadata)

    if response.status_code == 201:
        print(f"Folder {folder_name} is created successfully")
        return response.json()
    else:
        print(
            f"Failed to create folder {folder_name}, status code: {response.status_code}"
        )
        return None


def upload_folder_to_onedrive(access_token, local_folder_path, parent_folder_id):
    # Create the folder in OneDrive
    if len(parent_folder_id) < 20:
        parent_folder_id = get_folder_id(
            access_token, "FEI_SHARED/Repository/Bitbucket Repos"
        )

    folder_name = os.path.basename(local_folder_path)
    folder_metadata = create_folder_in_onedrive(
        access_token, parent_folder_id, folder_name
    )

    if folder_metadata is None:
        print("Failed to create folder, skipping this folder")
        return

    folder_id = folder_metadata.get("id")

    # Upload files in the folder recursively
    for item in os.listdir(local_folder_path):
        item_path = os.path.join(local_folder_path, item)
        if os.path.isfile(item_path):
            upload_file_to_onedrive(access_token, folder_id, item_path)
        elif os.path.isdir(item_path):
            upload_folder_to_onedrive(access_token, item_path, folder_id)


def rmtree(top):
    for root, dirs, files in os.walk(top, topdown=False):
        for name in files:
            filename = os.path.join(root, name)
            try:
                os.chmod(filename, stat.S_IWRITE)
                os.remove(filename)
            except PermissionError as e:
                print(
                    "UNABLE TO DELETE FILE: "
                    + filename
                    + "\n Something is locking the file...."
                )
                print(e.winerror)
        for name in dirs:
            os.rmdir(os.path.join(root, name))
    try:
        os.rmdir(top)
    except FileNotFoundError:
        pass


def change_permission(path):
    for root, dirs, files in os.walk(path, topdown=False):
        for name in files:
            filename = os.path.join(root, name)
            os.chmod(filename, stat.S_IWRITE)


def parse_excel():
    # read excel file
    wb = openpyxl.load_workbook(".\Config\Sync.xlsx")
    plant_sheets = []

    # sheet = wb.active
    print(wb.sheetnames)
    for sheet in wb.sheetnames:
        if sheet == "BitBucket Repos":
            repo_sheet = wb[sheet]
        elif sheet == "Sheet1":
            repo_sheet = wb[sheet]
        else:
            plant_sheets.append(wb[sheet])

    # Parse Repository Sheet First:
    for row in repo_sheet.iter_rows(values_only=True):
        row_data = []

        for cell in row:
            row_data.append(cell)

        if row_data[0] != "Name" and row_data[0] != None:
            repo_names.append(row_data[0])
        if row_data[1] != "Local Path" and row_data[1] != None:
            repo_staging_spots.append(row_data[1])
        if row_data[2] != "Destination Path" and row_data[2] != None:
            repo_onedrive_spots.append(row_data[2])

    # Parse Each Plant Sheet Next:
    for sheet in plant_sheets:
        for row in sheet.iter_rows(values_only=True):
            row_data = []

            for cell in row:
                row_data.append(cell)

            if row_data[0] != "Source" and row_data[0] != None:
                plant_onedrive_spots.append(row_data[0])
            if row_data[1] != "Destination" and row_data[1] != None:
                plant_server_spots.append(row_data[1])
            if row_data[2] != "Trigger":
                triggers.append(row_data[2])
            if row_data[3] != "Repository":
                repository.append(row_data[3])


def get_config():
    try:
        os.remove(".\Config\Sync.xlsx")
    except FileNotFoundError:
        print("getting config file...")
    access_token = get_access_token(client_id, client_secret, tenant)
    download_file_from_onedrive(
        access_token, "FEI_SHARED/Repository/Sync.xlsx", ".\Config\Sync.xlsx"
    )
    sleep(3)
    parse_excel()


# TODO TEST THIS FUNCTION:
def set_config():
    # Edit Excel and remove 'X' triggers
    wb = openpyxl.load_workbook("./Config/Sync.xlsx")
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet:
            for cell in row:
                if cell.value == "X":
                    cell.value = ""

    wb.save("./Config/Sync.xlsx")
    access_token = get_access_token(client_id, client_secret, tenant)
    folder_id = get_folder_id(access_token, "FEI_SHARED/Repository")
    upload_file_to_onedrive(access_token, folder_id, "./Config/Sync.xlsx")


def repo_sync(repo_name, path_to):
    # NOTE: path_to needs to be an empty directory
    # Delete directory if it exists
    try:
        rmtree(path_to)
        # rmtree(final_destination)
    except FileNotFoundError:
        print("Stage already empty")

    # Proxy works for this
    # NOTE repo_name must belong to fargoengineeringinc workspace
    git_link = f"https://{username}:{password}@bitbucket.org/fargoengineeringinc/{repo_name}.git"

    repo.clone_from(
        git_link, path_to, config="http.proxy=farft01:fardnc01@proxy-us.cnhind.com:8080"
    )

    # Append license to correct files
    with open("license.txt", "r") as l:
        license = l.read()
    for root, dirs, files in os.walk(path_to, topdown=True):
        for file in files:
            if file.endswith(".cs") or file.endswith(".ino") or file.endswith(".h"):
                try:
                    with open(os.path.join(root, file), "r", encoding="utf-8-sig") as f:
                        content = f.read()
                    with open(os.path.join(root, file), "wb") as f:
                        f.seek(0, 0)
                        new_content = (license + "\n\n" + format(content)).encode(
                            "utf-8"
                        )
                        f.write(new_content)
                except UnicodeDecodeError as e:
                    print("ADD LICENSE: Unable to parse target file: " + str(e))

    # rmtree(final_destination)
    rmtree(path_to + "\\.git")
    change_permission(path_to)
    access_token = get_access_token(client_id, client_secret, tenant)
    upload_folder_to_onedrive(
        access_token,
        path_to,
        repo_name,
    )


def server_sync(server_location, onedrive_folder):
    # Download OneDrive Folders specified in Excel
    # Depending on Trigger,
    # Move downloaded folders to server locations specified in excel.
    # Update Excel Triggers when complete
    print("Syncing to Server: ")

    path_to = "C:\\Temp\\Sync\\Docs"

    access_token = get_access_token(client_id, client_secret, tenant)

    download_folder_from_onedrive(access_token, onedrive_folder, path_to)

    rmtree(server_location)
    try:
        shutil.move(path_to, server_location)
    except Exception as e:
        print(e)


# Main Function
if __name__ == "__main__":
    start = datetime.now()
    get_config()

    print("CONFIGURATION RECEIVED")

    for i in range(len(repo_names)):
        if triggers[i] in ["W", "X"]:
            repo_sync(repo_names[i], f"{repo_staging_spots[i]}\\{repo_names[i]}")
        else:
            print(f"Ignoring Repo: {repo_names[i]}")

    # NEED TO MAKE THIS UNIQUE FOR EACH PLANT LOCATION/SHEET
    for i in range(len(plant_server_spots)):
        if triggers[i] in ["W", "X"]:
            server_sync(plant_server_spots[i], plant_onedrive_spots[i])
        else:
            print(f"Ignoring server path: {plant_server_spots[i]}")

    # if 'X' edit excel before we reupload it
    set_config()

    end = datetime.now()
    time = end - start
    print("DONE @ " + str(time))